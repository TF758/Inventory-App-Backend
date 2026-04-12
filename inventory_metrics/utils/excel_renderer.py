from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2



def render_workbook(spec: dict) -> Workbook:
    wb = Workbook()
    default_ws = wb.active
    created_any = False

    for sheet_name, sheet in spec.items():
        ws = wb.create_sheet(title=str(sheet_name)[:31])
        created_any = True

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])

        if headers:
            ws.append(headers)

        for row in rows:
            ws.append(row)

    # Remove default sheet if we created our own
    if created_any and default_ws.title == "Sheet":
        wb.remove(default_ws)

    # Safety: Excel REQUIRES at least one visible sheet
    if not wb.sheetnames:
        ws = wb.create_sheet(title="Report")
        ws.append(["No data available"])

    return wb

def render_workbook_streaming(spec: dict) -> Workbook:

    wb = Workbook(write_only=True)

    for sheet_name, sheet in spec.items():

        ws = wb.create_sheet(title=str(sheet_name)[:31])

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])

        if headers:
            ws.append(headers)

        # rows can now be generator OR list
        for row in rows:
            ws.append(row)

    return wb

def estimate_excel_size_mb(row_count: int, avg_row_bytes: int = 220) -> float:
    """
    Estimate XLSX file size for audit history exports.

    XLSX files compress well, but audit rows average
    roughly ~200–250 bytes once written.

    Returns MB estimate.
    """

    estimated_bytes = row_count * avg_row_bytes

    mb = estimated_bytes / (1024 * 1024)

    return round(mb, 2)