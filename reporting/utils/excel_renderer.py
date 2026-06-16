from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from django.utils.timezone import is_aware
from decimal import Decimal
from openpyxl.cell import WriteOnlyCell

CURRENCY_FORMAT = "$#,##0.00"

def autosize_columns(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            if cell.value is not None:
                max_length = max(
                    max_length,
                    len(str(cell.value))
                )
        ws.column_dimensions[col_letter].width = max_length + 2


def excel_safe(value):
    if isinstance(value, datetime) and is_aware(value):
        return value.replace(tzinfo=None)
    return value

def apply_sheet_formats(ws, headers, formats):
    """
    Apply workbook-spec formats to worksheet columns.
    """

    if not headers or not formats:
        return

    header_map = {
        header: idx + 1
        for idx, header in enumerate(headers)
    }

    for column_name, format_type in formats.items():

        column_index = header_map.get(column_name)

        if not column_index:
            continue

        for row_num in range(2, ws.max_row + 1):

            cell = ws.cell(
                row=row_num,
                column=column_index,
            )

            if cell.value is None:
                continue

            if format_type == "currency":

                if isinstance(
                    cell.value,
                    (int, float, Decimal),
                ):
                    cell.number_format = (
                        CURRENCY_FORMAT
                    )


def render_workbook(spec: dict) -> Workbook:
    wb = Workbook()

    default_ws = wb.active
    created_any = False

    for sheet_name, sheet in spec.items():

        ws = wb.create_sheet(
            title=str(sheet_name)[:31]
        )

        created_any = True

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])
        formats = sheet.get("formats", {})

        if headers:
            ws.append([
                excel_safe(v)
                for v in headers
            ])

        for row in rows:
            ws.append([
                excel_safe(v)
                for v in row
            ])

        apply_sheet_formats(
            ws,
            headers,
            formats,
        )

        autosize_columns(ws)

    if (
        created_any
        and default_ws.title == "Sheet"
    ):
        wb.remove(default_ws)

    if not wb.sheetnames:
        ws = wb.create_sheet(
            title="Report"
        )
        ws.append(
            ["No data available"]
        )

    return wb

def render_workbook_streaming(spec: dict) -> Workbook:

    wb = Workbook(write_only=True)

    for sheet_name, sheet in spec.items():

        ws = wb.create_sheet(
            title=str(sheet_name)[:31]
        )

        headers = sheet.get("headers", [])
        rows = sheet.get("rows", [])
        formats = sheet.get("formats", {})

        # -------------------------------------
        # Header mapping
        # -------------------------------------

        header_map = {
            header: idx
            for idx, header in enumerate(headers)
        }

        currency_columns = {
            header_map[column_name]
            for column_name, format_type in formats.items()
            if (
                format_type == "currency"
                and column_name in header_map
            )
        }

        # -------------------------------------
        # Headers
        # -------------------------------------

        if headers:
            ws.append(
                [excel_safe(v) for v in headers]
            )

        # -------------------------------------
        # Rows
        # -------------------------------------

        for row in rows:

            output_row = []

            for col_index, value in enumerate(row):

                value = excel_safe(value)

                if (
                    col_index in currency_columns
                    and isinstance(
                        value,
                        (int, float, Decimal),
                    )
                ):
                    cell = WriteOnlyCell(
                        ws,
                        value=value,
                    )

                    cell.number_format = (
                        CURRENCY_FORMAT
                    )

                    output_row.append(cell)

                else:
                    output_row.append(value)

            ws.append(output_row)

def estimate_excel_size_mb(row_count: int, avg_row_bytes: int = 220) -> float:
    """
    Estimate XLSX file size for audit history exports.

    XLSX files compress well, but audit rows average
    roughly ~200–250 bytes once written.

    Returns MB estimate.
    """

    estimated_bytes = row_count * avg_row_bytes

    mb = estimated_bytes / (1024 * 1024)

    return round(mb, 2)